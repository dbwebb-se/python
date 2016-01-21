"""
Python to Excel
================

This example scripts is written by Felix Khoi and will go over how you can write from Python to Excel.

The example will contain two versions, one simple and one more advanced.
In the simple one you will have to manually write code to write to Excel, it just contains a example of how you do it.
The more advanced script will show a user dialog where it lets the user select excel-file, date, hours and project.

OpenPyXL
===========

The script uses the import [OpenPyXL](https://openpyxl.readthedocs.org/en/default/).
"""


import os
import time

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

def main():
    """Main function of script"""

    intro()

    time_file = input("Enter file (don't forget .xlsx): ")
    date = input("Enter date - Day/Month/Year\n(write \"now\" to use todays date): ")
    hours = input("Enter worked hours: ")
    project = input("Enter project name: ")

    logToExcel(time_file, [date, hours, project])

def intro():
    """Intro to the script"""
    print("\nThis is a worked hours tracker, made by Felix Khoi.\n")
    print("This script will save given date, hours and project to given excel-file.")
    print("If the file doesn't exist, it will create a new excel-file.\n")

def logToExcel(time_file, worked_hours):
    """Function that logs to excel-file"""

    print("\nStarting to search for file...")

    #Checks if file already exists
    if os.path.isfile(time_file):
        print("\nFound " + time_file + "!")

        #Loading excel file
        time_tracker_sheet = load_workbook(filename=time_file)

        #Activates writing for given file
        writer = time_tracker_sheet.active
    else:
        #If file doesn't exist, create new excel-file
        time_tracker_sheet = Workbook()

        #Activates writer for new excel-file
        writer = time_tracker_sheet.active

        #Writes headers into new excel-file
        writer['A1'] = "Date"
        writer['B1'] = "Number of hours"
        writer['C1'] = "Project"

        #Adding bold and increasing font size
        writer['A1'].font = Font(bold=True, size=16)
        writer['B1'].font = Font(bold=True, size=16)
        writer['C1'].font = Font(bold=True, size=16)

    date = worked_hours[0]

    #Checks if user wants to use today's date
    if worked_hours[0] == "now":
        #Gets today's date
        date = time.strftime("%d/%m/%Y")

    #Appends worked hours to excel-file
    writer.append([date, worked_hours[1], worked_hours[2]])

    #Sets the width of the columns "A", "B", "C"
    writer.column_dimensions["A"].width = 50.0
    writer.column_dimensions["B"].width = 50.0
    writer.column_dimensions["C"].width = 50.0

    #Saves excel-file
    time_tracker_sheet.save(time_file)

    print("\nDone, saved to " + time_file + "\n")

if __name__ == "__main__":
    main()
