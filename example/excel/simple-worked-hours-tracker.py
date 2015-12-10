"""
Python to Excel
================

This example scripts is written by Felix Khoi and will go over how you can write from Python to Excel.

The example will contain two versions, one simple and one more advanced.
In the simple one you will have to manually write code to write to Excel, it just contains a example of how you do it.
The more advanced script will show a user dialog where it lets the user select excel-file, date, hours and project.

OpenPyXL
===========

The script uses the import [OpenPyXL](https://openpyxl.readthedocs.org/en/default/).
"""

import os
import time

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

def main():
    """Main function of script"""

    #Checks if file already exists
    if os.path.isfile("time-tracker.xlsx"):
        print("Found time tracker file!")

        #Loading excel file
        time_tracker_sheet = load_workbook(filename='time-tracker.xlsx')

        #Activates writing for given file
        writer = time_tracker_sheet.active
    else:
        #If file doesn't exist, create new excel-file
        time_tracker_sheet = Workbook()

        #Activates writer for new excel-file
        writer = time_tracker_sheet.active

        #Writes headers into new excel-file
        writer['A1'] = "Date"
        writer['B1'] = "Number of hours"
        writer['C1'] = "Project"

        #Adding bold and increasing font size
        writer['A1'].font = Font(bold=True, size=16)
        writer['B1'].font = Font(bold=True, size=16)
        writer['C1'].font = Font(bold=True, size=16)

    #Gets today's date
    dateAndTime = time.strftime("%d/%m/%Y")

    #Appends worked hours to excel-file
    writer.append([dateAndTime, 3, "Python Project 1"])
    writer.append([dateAndTime, 7, "Python Project 3"])
    writer.append(["03/12/2015", 6, "Python Project 2"])
    writer.append(["05/12/2015", 1, "Python Project 4"])

    #Sets the width of the columns "A", "B", "C"
    writer.column_dimensions["A"].width = 50.0
    writer.column_dimensions["B"].width = 50.0
    writer.column_dimensions["C"].width = 50.0

    #Saves excel-file
    time_tracker_sheet.save("time-tracker.xlsx")

    print("Done, saved to time-tracker.xlsx")

if __name__ == "__main__":
    main()
